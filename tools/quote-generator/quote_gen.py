#!/usr/bin/env python3
"""
quote_gen.py — YAKUN 报价表生成器
根据客户清单（SKU + 数量）自动生成正式报价单 xlsx。

用法:
    python3 quote_gen.py 清单.txt [-o 输出.xlsx] [-d 2026-09-02]

清单文件格式（每行一条，支持逗号/制表符/空格分隔）:
    ZQS-1440, 1
    MSP-S01 2
    PG-WHT-01	3

依赖: openpyxl（pip install openpyxl）
"""
import sys, os, json, argparse
from datetime import date

SKU_MAP = {
    'ZQS-1440': 'ZQS1440', 'ZQS-8149': 'ZQS8149',
    'ZQS-14102': 'ZQS14102', 'ZQS-14103': 'ZQS14103',
}

def load_products():
    base = os.path.dirname(os.path.abspath(__file__))
    data = json.load(open(os.path.join(base, '..', '..', 'data', 'products.json'), encoding='utf-8'))
    prods = {}
    for p in data['products']:
        prods[p['sku']] = p
        prods[p['sku'].replace('-', '')] = p  # 别名：去连字符
        for v in (p.get('variants') or []):
            prods[v['sku']] = {**p, 'price': v.get('price', 0), 'name': v.get('label', p['name'])}
            prods[v['sku'].replace('-', '')] = prods[v['sku']]
    return prods

def parse_list(path):
    items = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = [x for x in line.replace('\t', ',').replace('，', ',').split(',') if x.strip()]
            if len(parts) >= 2:
                sku, qty = parts[0].strip(), int(float(parts[1].strip()))
            else:  # 空格分隔
                sku, qty = parts[0].split()[:2]
                qty = int(qty)
            items.append((sku, qty))
    return items

def gen(quote_date, items, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    prods = load_products()
    wb = Workbook(); ws = wb.active; ws.title = 'Quotation'
    navy, orange, light = '0B2447', 'FF6B1A', 'F5F7FB'
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:G1'); ws['A1'] = 'PT YAKUN PRIMA INDONESIA'
    ws['A1'].font = Font(size=16, bold=True, color=navy)
    ws.merge_cells('A2:G2'); ws['A2'] = f'QUOTATION · 报价单 · 报价日期 {quote_date} · 有效期 7 天'
    ws['A2'].font = Font(size=9, color='64748B')
    ws.merge_cells('A3:G3'); ws['A3'] = 'WhatsApp: +62 821-1876-584 · Gudang: Jakarta, Indonesia'
    ws['A3'].font = Font(size=9, color='64748B')

    for i, h in enumerate(['NO', 'SKU', 'Nama Produk (ID)', '产品名称 (ZH)', 'Jumlah (PCS)', 'Harga Satuan (IDR)', 'Subtotal (IDR)'], 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=navy)
        c.alignment = center; c.border = border

    total_priced = total_qty = 0
    missing = []
    r = 6
    for idx, (sku_raw, qty) in enumerate(items, 1):
        p = prods.get(sku_raw) or prods.get(SKU_MAP.get(sku_raw, ''))
        if not p:
            missing.append(sku_raw)
            name_id = name_zh = sku_raw; price = 0
        else:
            name_id = p['name']['id']; name_zh = p['name']['zh']; price = p.get('price', 0)
        subtotal = price * qty if price else 0
        if price: total_priced += subtotal
        total_qty += qty
        vals = [idx, sku_raw, name_id, name_zh, qty, price or None, subtotal or None]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(size=10); c.border = border
            if i in (1, 2, 5): c.alignment = center
            elif i in (6, 7):
                c.alignment = right
                if v is not None: c.number_format = '#,##0'
        if r % 2 == 0:
            for i in range(1, 8): ws.cell(row=r, column=i).fill = PatternFill('solid', fgColor=light)
        r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value='TOTAL（已定价部分小计）').font = Font(size=10, bold=True, color=navy)
    ws.cell(row=r, column=5, value=total_qty).font = Font(size=10, bold=True)
    ws.cell(row=r, column=7, value=total_priced).font = Font(size=10, bold=True, color=orange)
    ws.cell(row=r, column=7).number_format = '#,##0'
    for i in range(1, 8):
        ws.cell(row=r, column=i).border = border
        ws.cell(row=r, column=i).fill = PatternFill('solid', fgColor='FFF3E8')
        ws.cell(row=r, column=i).alignment = center if i == 5 else (right if i == 7 else Alignment(vertical='center'))

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value='备注: 留空价格的产品请 WhatsApp 咨询。已定价价格为批发参考价（不含运费），最终以确认单为准。').font = Font(size=9, italic=True, color='94A3B8')

    for i, w in enumerate([5, 13, 38, 22, 12, 17, 17], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f'✅ 报价表: {out_path}')
    print(f'   共 {len(items)} 行 | 数量合计 {total_qty} PCS | 已定价小计 Rp {total_priced:,}')
    if missing:
        print(f'   ⚠️ 数据中找不到: {", ".join(missing)}')

if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='YAKUN 报价表生成器')
    ap.add_argument('list', help='客户清单文件（每行: SKU 数量）')
    ap.add_argument('-o', '--out', default=None, help='输出 xlsx 路径')
    ap.add_argument('-d', '--date', default=date.today().isoformat(), help='报价日期 YYYY-MM-DD')
    args = ap.parse_args()
    out = args.out or f'YAKUN-报价表-{args.date}.xlsx'
    items = parse_list(args.list)
    if not items:
        print('清单为空，请检查文件格式'); sys.exit(1)
    gen(args.date, items, out)
